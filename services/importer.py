"""
Excel spreadsheet importer — reads a job tracking spreadsheet
and cleans up typos during import.
"""

import os
from datetime import datetime
from openpyxl import load_workbook

# Typo corrections by field
CORRECTIONS = {
    "source": {
        "indede": "Indeed",
        "indded": "Indeed",
        "in": "Indeed",
        "indeed": "Indeed",
        "f-jobs": "FlexJobs",
        "stie": "Company Site",
        "site": "Company Site",
        "s": "Company Site",
        "amazon.jobs": "Amazon Jobs",
        "loves site": "Company Site",
        "nrm site": "Company Site",
        "email": "Email",
    },
    "status": {
        "appiked": "Applied",
        "applied.": "Applied",
        "applied ": "Applied",
        "applied": "Applied",
        "rejected": "Rejected",
    },
    "industry": {
        "sattelite": "Satellite",
        "manufactureing": "Manufacturing",
        "warehous": "Warehouse",
        "carr": "Automotive",
        "cars": "Automotive",
        "ware": "Warehouse",
        "pharm": "Pharma",
        "it suppor": "IT Support",
        "it support": "IT Support",
        "wastewater": "Wastewater",
        "waste": "Waste Management",
        "warehouse": "Warehouse",
        "roofing": "Roofing",
        "hospital": "Healthcare",
        "retail": "Retail",
        "logistics": "Logistics",
        "oil": "Oil & Gas",
        "wind": "Wind Energy",
    },
    "resume_used": {
        "core": "Core",
        "new": "New",
        "new ": "New",
        "it": "IT",
        "customer s": "Customer Service",
        "amazon tailored": "Amazon Tailored",
    },
    "company": {
        "mosaic receuitying": "Mosaic Recruiting",
        "amazon - stie": "Amazon",
        "carvana- indeed": "Carvana",
        "carvana - site": "Carvana",
        "communtiy hosp.": "Community Hospital",
        "tech partmers llf": "Tech Partners LLC",
        "titan porcesional sercies": "Titan Professional Services",
        "clea capital cubicas": "Clear Capital Cubics",
        "norman regional ho": "Norman Regional Hospital",
        "gxo - site": "GXO Logistics",
    },
    "title": {
        "warehouse wuperviso": "Warehouse Supervisor",
        "plant man": "Plant Manager",
        "reconditinng manager": "Reconditioning Manager",
        "ware house": "Warehouse Associate",
        "warehouse 2": "Warehouse Associate",
    },
}


def clean_value(field: str, raw: str) -> str:
    """Apply field-specific corrections to a raw value."""
    if not raw:
        return raw
    lookup = raw.strip().lower()
    corrections = CORRECTIONS.get(field, {})
    return corrections.get(lookup, raw.strip())


def import_spreadsheet(filepath: str) -> list[dict]:
    """Read the Excel spreadsheet, clean up data, return list of job dicts."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Spreadsheet not found: {filepath}")

    wb = load_workbook(filepath)
    ws = wb.active
    jobs = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = [cell.value for cell in row]

        # Skip empty rows and header rows
        if not any(cells):
            continue

        # Column layout: A=number, B=date, C=company, D=role, E=industry,
        #                F=source, G=resume, H=status
        row_num = cells[0]
        date_val = cells[1]
        company = cells[2]
        title = cells[3]
        industry = cells[4]
        source = cells[5]
        resume = cells[6]
        status = cells[7]

        # Skip rows without both company and title
        if not company and not title:
            continue
        # Skip if it looks like a header or category note (no date)
        if not date_val and not row_num:
            continue

        # Handle row 16 where company and title are swapped
        # (Ticketing Coordinator is in company column, Select Water Solutions in title)
        if company and title:
            company_str = str(company).strip()
            title_str = str(title).strip()
            if company_str == "Ticketing Coordinator":
                company, title = title, company
            elif title_str == "Select Water Soloutions":
                company, title = title, company

        # Parse date
        applied_date = None
        if date_val:
            if isinstance(date_val, datetime):
                applied_date = date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, str):
                applied_date = date_val

        # Clean values
        company_clean = clean_value("company", str(company)) if company else None
        title_clean = clean_value("title", str(title)) if title else None
        industry_clean = clean_value("industry", str(industry)) if industry else None
        source_clean = clean_value("source", str(source)) if source else "Unknown"
        resume_clean = clean_value("resume_used", str(resume)) if resume else None
        status_clean = clean_value("status", str(status)) if status else "Applied"

        # Map statuses to our schema
        status_map = {
            "Applied": "applied",
            "Rejected": "rejected",
            "Interviewing": "interviewing",
            "Offer": "offer",
            "New": "new",
        }
        final_status = status_map.get(status_clean, "applied")

        jobs.append({
            "source": "import",
            "title": title_clean or "Unknown",
            "company": company_clean,
            "industry": industry_clean,
            "location": "",  # Default if not specified in spreadsheet
            "resume_used": resume_clean,
            "status": final_status,
            "applied_date": applied_date,
            "created_at": applied_date or datetime.now().isoformat(),
            "notes": f"Imported from spreadsheet. Original source: {source_clean}",
        })

    return jobs


def insert_imported_jobs(conn, jobs: list[dict]) -> dict:
    """Insert imported jobs into database. Returns summary."""
    from services.db import upsert_job

    inserted = 0
    skipped = 0
    for job in jobs:
        # Use a synthetic URL for dedup on reimport
        job["url"] = f"import://{job.get('company', 'unknown')}/{job.get('title', 'unknown')}/{job.get('applied_date', 'unknown')}"
        result = upsert_job(conn, job)
        if result > 0:
            inserted += 1
        else:
            skipped += 1

    return {"inserted": inserted, "skipped": skipped, "total": len(jobs)}
